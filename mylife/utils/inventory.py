from openpyxl import Workbook, load_workbook

def product_search(warehouse, ws1, ws2, ws3, ws4):
    i1 = 2
    invent = []
    warehouse = warehouse
    ws1 = ws1
    ws2 = ws2
    ws3 = ws3
    ws4 = ws4
    while(True):
        code = ws1.cell(i1,3).value
        if code == None:
            break

        number = section_search(code, warehouse, ws1, ws2, ws3, ws4)
        name = ws1.cell(i1, 2).value

        # 更新铂金库存
        if code=='PTXYA3536':
            max = 0
            max = bojin_search(code, warehouse, ws1, ws2, ws3, ws4)
            # 一瓶巴拿马 等于 20瓶麦克斯铂金
            number = number + max*20

        # 更新黑金库存
        if code=='PTXZA0065':
            max = 0
            max = bojin_search(code, warehouse, ws1, ws2, ws3, ws4)
            # 一瓶五粮液八代XMT 等于 4瓶麦克斯黑金
            number = number + max*4

        invent.append([name, code, number])
        i1 = i1 +1
        # print('prodect get:')
        # print(number)
    return(invent)


def section_search(code, warehouse, ws1, ws2, ws3, ws4):
    i2 = 2
    warehouse = warehouse
    ws1 = ws1
    ws2 = ws2
    ws3 = ws3
    ws4 = ws4
    while(True):
        section = ws2.cell(i2,1).value
        if section == None:
            break
        for x in range(3,7):
            section = ws2.cell(i2, x).value
            if section == code:
                number = 0
                for x in range(3,7):
                    check = ws2.cell(i2, x).value
                    if check != None:
                        x = erp_search(check, warehouse, ws1, ws2, ws3, ws4)
                        number = number + x
                # print('section return:')
                # print(number)
                return(number)

        i2 = i2 + 1


def erp_search(check, warehouse, ws1, ws2, ws3, ws4):
    i3 = 2
    number = 0
    warehouse = warehouse
    ws1 = ws1
    ws2 = ws2
    ws3 = ws3
    ws4 = ws4
    while(True):
        erp_section = ws3.cell(i3, 3).value
        if erp_section == None:
            break
        print('check is :'+check)
        print(erp_section)
        if erp_section == check and ws3.cell(i3, 1).value == warehouse:
            x = ws3.cell(i3, 9).value
            number = number + int(x)
            print(x)
        i3 = i3 + 1
        # print('erp return:')
        # print(number)
    return(number)

# 在第二张表搜索铂金库存
def bojin_search(code, warehouse, ws1, ws2, ws3, ws4):
    i = 3
    max = 0
    ws1 = ws1
    ws2 = ws2
    ws3 = ws3
    ws4 = ws4
    while(True):
        if ws4.cell(i, 2).value == warehouse and ws4.cell(i, 3).value == 'BJ18A0208':
            max = ws4.cell(i, 8).value
            return(max)
        if warehouse == '武汉总仓':
            if ws4.cell(i, 2).value == '武东分仓' and ws4.cell(i, 3).value == 'BJ18A0208':
                max = ws4.cell(i, 8).value
                return(max)
        if ws4.cell(i, 2).value == None:
            break
        i = i + 1
    return(max)
  
# 在第二张表搜索黑金库存
def bojin_search(code, warehouse, ws1, ws2, ws3, ws4):
    i = 3
    max = 0
    ws1 = ws1
    ws2 = ws2
    ws3 = ws3
    ws4 = ws4
    while(True):
        if ws4.cell(i, 2).value == warehouse and ws4.cell(i, 3).value == 'BJ21A0128':
            max = ws4.cell(i, 8).value
            return(max)
        if warehouse == '武汉总仓':
            if ws4.cell(i, 2).value == '武东分仓' and ws4.cell(i, 3).value == 'BJ21A0128':
                max = ws4.cell(i, 8).value
                return(max)
        if ws4.cell(i, 2).value == None:
            break
        i = i + 1
    return(max)

def inventory(name, danpin, kuanhao, kucun):
    wb1 = load_workbook(danpin)
    wb2 = load_workbook(kuanhao)
    wb3 = load_workbook(kucun)

    ws1 = wb1.active
    ws2 = wb2.active
    ws3 = wb3.active
    ws4 = wb3.worksheets[1] # 抓取奔富麦克斯库存表格 

    name = name
    invent = product_search(name, ws1, ws2, ws3, ws4)

    return invent