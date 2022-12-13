from django.shortcuts import render
from django.http import HttpResponse, JsonResponse

from openpyxl import Workbook, load_workbook
from tempfile import NamedTemporaryFile

from urllib import response

# Create your views here.
def index(request):
    return HttpResponse('Hello world')

def test(request):
    return render(request, 'index/test.html')

def ajax_login(request):
    wb = Workbook()
    ws = wb.active

    ws.cell(1,1).value = 'ok'

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
    response = HttpResponse(content=stream, content_type='application/ms-excel', )
    
    response['Content-Disposition'] = f'attachment; filename=test.xlsx'
    return response