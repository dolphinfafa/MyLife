from urllib import response
import webbrowser
from django.shortcuts import render
from django.views.generic import View

from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from tempfile import NamedTemporaryFile

from utils.inventory import inventory

# Create your views here.
class InventView(View):
    def get(self, request):
        return render(request, 'yesmywine/invent.html')

    def post(self, request):
        danpin = request.FILES.get('danpin')
        kuanhao = request.FILES.get('kuanhao')
        kucun = request.FILES.get('kucun')
        try:
            if danpin is None:
                return HttpResponse('请选择要上传的文件')
            # 循环二进制写入
            wb1 = load_workbook(danpin)
            wb2 = load_workbook(kuanhao)
            wb3 = load_workbook(kucun)

            baishazhou = '武汉总仓'
            wansongyuan = '经济万松园店'
            ezhou = '仓储鄂州阳光货仓店'
            fuxin = '仓储富鑫常青店'
            sifang = '仓储肆方光谷店'
            binjiang = '仓储滨江国际店'

            if 'baishazhou' in request.POST:
                name = baishazhou
            elif 'wansongyuan' in request.POST:
                name = wansongyuan
            elif 'ezhou' in request.POST:
                name = ezhou
            elif 'fuxin' in request.POST:
                name = fuxin
            elif 'sifang' in request.POST:
                name = sifang
            elif 'binjiang' in request.POST:
                name = binjiang

            invent = inventory(name, danpin, kuanhao, kucun)

            wb = Workbook()
            ws = wb.active

            x = 0
            for i in invent:
                ws.cell(x+2, 1).value = invent[x][0]
                ws.cell(x+2, 2).value = invent[x][1]
                ws.cell(x+2, 3).value = invent[x][2]
                x = x + 1

            with NamedTemporaryFile() as tmp:
                wb.save(tmp.name)
                tmp.seek(0)
                stream = tmp.read()
            response = HttpResponse(content=stream, content_type='application/ms-excel', )
            
            if name == baishazhou:
                name = 'baishazhou'
            elif name == wansongyuan:
                name = 'wansongyuan'
            elif name == ezhou:
                name = 'ezhou'
            elif name == fuxin:
                name = 'fuxin'
            elif name == sifang:
                name = 'sifang'
            elif name == binjiang:
                name = 'binjiang'
            
            s = f'attachment; filename={name}.xlsx'
            print(s)

            response['Content-Disposition'] = f'attachment; filename={name}.xlsx'
            return response
        except Exception as e:
            return HttpResponse(e)